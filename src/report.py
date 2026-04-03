import io
from datetime import datetime

import pandas as pd
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# --- Brand colors ---
_BRAND_GREEN_HEX = "4A7C28"
_BRAND_GREEN_DARK_HEX = "1B5E20"
_BRAND_GREEN_LIGHT_HEX = "E8F5E9"
_BRAND_ORANGE_HEX = "E65100"
_BRAND_ORANGE_LIGHT_HEX = "FFF3E0"

# Excel styles
HEADER_FILL = PatternFill(start_color=_BRAND_GREEN_HEX, end_color=_BRAND_GREEN_HEX, fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
BOLD_FONT = Font(bold=True, size=10)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
GREEN_FILL = PatternFill(start_color=_BRAND_GREEN_LIGHT_HEX, end_color=_BRAND_GREEN_LIGHT_HEX, fill_type="solid")
ORANGE_FILL = PatternFill(start_color=_BRAND_ORANGE_LIGHT_HEX, end_color=_BRAND_ORANGE_LIGHT_HEX, fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
STATUS_FILLS = {"OK": GREEN_FILL, "UPOZORENJE": YELLOW_FILL, "GREŠKA": ORANGE_FILL}


def _write_header_row(ws, row_num: int, headers: list[str], widths: list[int] | None = None):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _auto_width(ws, min_width=8, max_width=30):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


# ═══════════════════════════════════════════════════
# EXCEL REPORT
# ═══════════════════════════════════════════════════

def generate_excel_report(
    summary_df: pd.DataFrame,
    excel_df: pd.DataFrame,
    pdf_df: pd.DataFrame,
    unmatched_bank: pd.DataFrame,
    excel_meta: dict,
    pdf_meta: dict,
    match_details_func,
) -> bytes:
    wb = Workbook()

    # === Sheet 1: Pregled ===
    ws = wb.active
    ws.title = "Pregled"
    ws.sheet_properties.tabColor = _BRAND_GREEN_HEX

    period = f"{excel_meta['period_start'].strftime('%d.%m.%Y')} - {excel_meta['period_end'].strftime('%d.%m.%Y')}"

    ws.merge_cells("A1:G1")
    ws.cell(row=1, column=1, value="Tvornica Zdrave Hrane — Usporedba Prometa vs Banka").font = Font(bold=True, size=13, color=_BRAND_GREEN_DARK_HEX)
    ws.cell(row=2, column=1, value=f"Razdoblje: {period}").font = Font(size=10, color="666666")
    ws.cell(row=3, column=1, value=f"Generirano: {datetime.now().strftime('%d.%m.%Y %H:%M')}").font = Font(size=9, color="999999")

    headers = ["PJ", "Poslovnica", "Promet (€)", "Banka (€)", "Razlika (€)", "%", "Status"]
    _write_header_row(ws, 5, headers)

    for i, (_, row) in enumerate(summary_df.iterrows(), 6):
        ws.cell(row=i, column=1, value=row["pj_code"]).border = THIN_BORDER
        ws.cell(row=i, column=2, value=row["pj_name"]).border = THIN_BORDER
        for c, key in [(3, "promet_total"), (4, "banka_total"), (5, "razlika")]:
            cell = ws.cell(row=i, column=c, value=row[key])
            cell.border = THIN_BORDER
            cell.number_format = '#,##0.00'
        ws.cell(row=i, column=6, value=row["razlika_pct"]).border = THIN_BORDER
        ws.cell(row=i, column=6).number_format = '+0.0%;-0.0%'
        ws.cell(row=i, column=7, value=row["status"]).border = THIN_BORDER
        fill = STATUS_FILLS.get(row["status"], PatternFill())
        for c in range(1, 8):
            ws.cell(row=i, column=c).fill = fill

    tr = 6 + len(summary_df)
    ws.cell(row=tr, column=1, value="UKUPNO").font = BOLD_FONT
    for c, val in [(3, summary_df["promet_total"].sum()), (4, summary_df["banka_total"].sum()), (5, summary_df["razlika"].sum())]:
        cell = ws.cell(row=tr, column=c, value=val)
        cell.font = BOLD_FONT
        cell.number_format = '#,##0.00'
    for c in range(1, 8):
        ws.cell(row=tr, column=c).border = THIN_BORDER
    _auto_width(ws)

    # === Sheet 2: Detalji po poslovnici ===
    ws2 = wb.create_sheet("Detalji")
    ws2.sheet_properties.tabColor = _BRAND_GREEN_HEX
    OKSTAR_FILL = PatternFill(start_color="E0F2F1", end_color="E0F2F1", fill_type="solid")
    all_status_fills = {**STATUS_FILLS, "OK*": OKSTAR_FILL}

    cr = 1
    for _, srow in summary_df.iterrows():
        pj_code, pj_name = srow["pj_code"], srow["pj_name"]
        ws2.merge_cells(start_row=cr, start_column=1, end_row=cr, end_column=6)
        hcell = ws2.cell(row=cr, column=1, value=f"{pj_code} — {pj_name}")
        hcell.font = Font(bold=True, size=11, color=_BRAND_GREEN_DARK_HEX)
        cr += 1
        ws2.cell(row=cr, column=1, value=f"Promet: {srow['promet_total']:,.2f} €").font = Font(size=9, color="666666")
        ws2.cell(row=cr, column=3, value=f"Banka: {srow['banka_total']:,.2f} €").font = Font(size=9, color="666666")
        ws2.cell(row=cr, column=5, value=f"Razlika: {srow['razlika']:+,.2f} €").font = Font(size=9, color=_BRAND_ORANGE_HEX if abs(srow['razlika']) >= 5 else "666666")
        cr += 1
        _write_header_row(ws2, cr, ["Datum", "Promet (€)", "Banka (€)", "Razlika (€)", "Status", "Napomena"])
        cr += 1
        details = match_details_func(excel_df, pdf_df, pj_code)
        for _, d in details.iterrows():
            ws2.cell(row=cr, column=1, value=d["date"].strftime("%d.%m.%Y") if pd.notna(d["date"]) else "").border = THIN_BORDER
            for c, key in [(2, "promet"), (3, "banka"), (4, "razlika")]:
                cell = ws2.cell(row=cr, column=c, value=d[key])
                cell.border = THIN_BORDER
                cell.number_format = '#,##0.00'
            ws2.cell(row=cr, column=5, value=d["status"]).border = THIN_BORDER
            napomena = d.get("napomena", "")
            ws2.cell(row=cr, column=6, value=napomena).border = THIN_BORDER
            ws2.cell(row=cr, column=6).font = Font(size=8, color="888888")
            fill = all_status_fills.get(d["status"], PatternFill())
            for c in range(1, 7):
                ws2.cell(row=cr, column=c).fill = fill
            cr += 1
        cr += 1
    _auto_width(ws2)

    # === Sheet 3: Neuparene ===
    ws3 = wb.create_sheet("Neuparene")
    ws3.cell(row=1, column=1, value="Neuparene bankovne transakcije").font = Font(bold=True, size=11, color=_BRAND_GREEN_DARK_HEX)
    if not unmatched_bank.empty:
        _write_header_row(ws3, 3, ["Rb", "Datum", "Iznos (€)", "PNB", "Opis"])
        for i, (_, row) in enumerate(unmatched_bank.iterrows(), 4):
            ws3.cell(row=i, column=1, value=row["row_num"]).border = THIN_BORDER
            ws3.cell(row=i, column=2, value=row["booking_date"].strftime("%d.%m.%Y") if pd.notna(row["booking_date"]) else "").border = THIN_BORDER
            ws3.cell(row=i, column=3, value=row["amount"]).border = THIN_BORDER
            ws3.cell(row=i, column=3).number_format = '#,##0.00'
            ws3.cell(row=i, column=4, value=row["pnb"]).border = THIN_BORDER
            ws3.cell(row=i, column=5, value=row["description"]).border = THIN_BORDER
    else:
        ws3.cell(row=3, column=1, value="Nema neuparenih stavki.")
    _auto_width(ws3)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═══════════════════════════════════════════════════
# PDF REPORT
# ═══════════════════════════════════════════════════

_FONT_NAME = "Helvetica"

# Brand RGB tuples
_GREEN = (74, 124, 40)
_GREEN_DARK = (27, 94, 32)
_GREEN_LIGHT = (232, 245, 233)
_ORANGE = (230, 81, 0)
_ORANGE_LIGHT = (255, 243, 224)
_GRAY = (100, 100, 100)
_GRAY_LIGHT = (245, 245, 240)
_WHITE = (255, 255, 255)


class PDFReport(FPDF):

    def __init__(self, period_str="", **kwargs):
        super().__init__(**kwargs)
        self._period = period_str
        # Helvetica is built into fpdf2, no add_font needed

    def header(self):
        # Green top bar
        self.set_fill_color(*_GREEN)
        self.rect(0, 0, self.w, 3, "F")
        # Title
        self.set_y(8)
        self.set_font(_FONT_NAME, "B", 13)
        self.set_text_color(*_GREEN_DARK)
        self.cell(0, 7, "Tvornica Zdrave Hrane", new_x="LMARGIN", new_y="NEXT")
        self.set_font(_FONT_NAME, "", 9)
        self.set_text_color(*_GRAY)
        self.cell(0, 5, f"Usporedba prometa i bankovnih izvoda  |  {self._period}", new_x="LMARGIN", new_y="NEXT")
        self.ln(3)

    def footer(self):
        self.set_y(-12)
        self.set_font(_FONT_NAME, "", 7)
        self.set_text_color(180, 180, 180)
        self.cell(0, 8, f"Generirano: {datetime.now().strftime('%d.%m.%Y %H:%M')}  |  Stranica {self.page_no()}/{{nb}}", align="C")

    def _table_header(self, widths, headers):
        self.set_font(_FONT_NAME, "B", 8)
        self.set_fill_color(*_GREEN)
        self.set_text_color(*_WHITE)
        for w, h in zip(widths, headers):
            self.cell(w, 6, h, border=1, fill=True, align="C")
        self.ln()
        self.set_text_color(0, 0, 0)

    def _status_fill(self, status):
        if status == "OK":
            self.set_fill_color(*_GREEN_LIGHT)
        elif status == "OK*":
            self.set_fill_color(224, 242, 241)
        elif status == "UPOZORENJE":
            self.set_fill_color(255, 248, 225)
        else:
            self.set_fill_color(*_ORANGE_LIGHT)

    def _section_title(self, title, subtitle=""):
        self.set_font(_FONT_NAME, "B", 10)
        self.set_text_color(*_GREEN_DARK)
        # Green dot
        y = self.get_y() + 2
        self.set_fill_color(*_GREEN)
        self.ellipse(self.l_margin, y, 3, 3, "F")
        self.set_x(self.l_margin + 5)
        self.cell(0, 7, title, new_x="LMARGIN", new_y="NEXT")
        if subtitle:
            self.set_font(_FONT_NAME, "", 8)
            self.set_text_color(*_GRAY)
            self.cell(0, 4, subtitle, new_x="LMARGIN", new_y="NEXT")
        self.ln(2)


def generate_pdf_report(
    summary_df: pd.DataFrame,
    excel_df: pd.DataFrame,
    pdf_df: pd.DataFrame,
    excel_meta: dict,
    pdf_meta: dict,
    match_details_func,
) -> bytes:
    period = f"{excel_meta['period_start'].strftime('%d.%m.%Y')} — {excel_meta['period_end'].strftime('%d.%m.%Y')}"
    pdf = PDFReport(period_str=period, orientation="P")
    pdf.alias_nb_pages()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=18)

    pw = pdf.w - pdf.l_margin - pdf.r_margin  # usable page width

    # --- KPI boxes ---
    promet_total = summary_df["promet_total"].sum()
    banka_total = summary_df["banka_total"].sum()
    razlika_total = summary_df["razlika"].sum()
    ok_count = len(summary_df[summary_df["status"] == "OK"])
    err_count = len(summary_df[summary_df["status"] != "OK"])

    box_w = pw / 4 - 1.5
    box_h = 14
    y0 = pdf.get_y()

    kpis = [
        ("Promet", f"{promet_total:,.2f} €", _GREEN_LIGHT, _GREEN_DARK),
        ("Banka", f"{banka_total:,.2f} €", _GREEN_LIGHT, _GREEN_DARK),
        ("Razlika", f"{razlika_total:+,.2f} €", _ORANGE_LIGHT, _ORANGE),
        ("Status", f"{ok_count} OK / {err_count} razl.", _GRAY_LIGHT, (60, 60, 60)),
    ]
    for i, (label, value, bg, fg) in enumerate(kpis):
        x = pdf.l_margin + i * (box_w + 2)
        pdf.set_fill_color(*bg)
        pdf.rect(x, y0, box_w, box_h, "DF")
        pdf.set_xy(x + 3, y0 + 1)
        pdf.set_font(_FONT_NAME, "", 6.5)
        pdf.set_text_color(*_GRAY)
        pdf.cell(box_w - 6, 4, label.upper())
        pdf.set_xy(x + 3, y0 + 5.5)
        pdf.set_font(_FONT_NAME, "B", 10)
        pdf.set_text_color(*fg)
        pdf.cell(box_w - 6, 6, value)

    pdf.set_y(y0 + box_h + 6)

    # --- Summary table ---
    pdf._section_title("Pregled po poslovnicama")

    widths = [14, 38, 28, 28, 28, 16, 18]
    scale = pw / sum(widths)
    widths = [w * scale for w in widths]
    headers = ["PJ", "Poslovnica", "Promet (€)", "Banka (€)", "Razlika (€)", "%", "Status"]
    pdf._table_header(widths, headers)

    pdf.set_font(_FONT_NAME, "", 7.5)
    rh = 5.5
    for _, row in summary_df.iterrows():
        if pdf.get_y() + rh > pdf.h - 20:
            pdf.add_page()
            pdf._table_header(widths, headers)
            pdf.set_font(_FONT_NAME, "", 7.5)

        pdf._status_fill(row["status"])
        pdf.cell(widths[0], rh, str(row["pj_code"]), border=1, fill=True)
        pdf.cell(widths[1], rh, str(row["pj_name"])[:22], border=1, fill=True)
        pdf.cell(widths[2], rh, f"{row['promet_total']:,.2f}", border=1, fill=True, align="R")
        pdf.cell(widths[3], rh, f"{row['banka_total']:,.2f}", border=1, fill=True, align="R")
        pdf.cell(widths[4], rh, f"{row['razlika']:+,.2f}", border=1, fill=True, align="R")
        pdf.cell(widths[5], rh, f"{row['razlika_pct']:+.1f}%", border=1, fill=True, align="R")
        pdf.cell(widths[6], rh, row["status"], border=1, fill=True, align="C")
        pdf.ln()

    # Totals row
    pdf.set_font(_FONT_NAME, "B", 7.5)
    pdf.set_fill_color(230, 230, 225)
    pdf.cell(widths[0] + widths[1], rh, "UKUPNO", border=1, fill=True)
    pdf.cell(widths[2], rh, f"{promet_total:,.2f}", border=1, fill=True, align="R")
    pdf.cell(widths[3], rh, f"{banka_total:,.2f}", border=1, fill=True, align="R")
    pdf.cell(widths[4], rh, f"{razlika_total:+,.2f}", border=1, fill=True, align="R")
    pdf.cell(widths[5] + widths[6], rh, "", border=1, fill=True)
    pdf.ln()

    # --- Detail tables per branch ---
    d_widths_raw = [20, 25, 25, 25, 14]
    d_scale = (pw * 0.65) / sum(d_widths_raw)
    d_widths = [w * d_scale for w in d_widths_raw]
    note_w = pw - sum(d_widths)  # remaining width for napomena
    d_headers = ["Datum", "Promet (€)", "Banka (€)", "Razlika (€)", "Status"]
    d_rh = 5

    for _, srow in summary_df.iterrows():
        pj_code, pj_name = srow["pj_code"], srow["pj_name"]
        details = match_details_func(excel_df, pdf_df, pj_code)
        has_notes = details["napomena"].str.len().sum() > 0

        needed = 20 + len(details) * d_rh + 10
        if pdf.get_y() + min(needed, 60) > pdf.h - 20:
            pdf.add_page()

        pdf._section_title(
            f"{pj_code} — {pj_name}",
            f"Promet: {srow['promet_total']:,.2f} €   |   Banka: {srow['banka_total']:,.2f} €   |   Razlika: {srow['razlika']:+,.2f} €",
        )

        all_headers = d_headers + (["Napomena"] if has_notes else [])
        all_widths = d_widths + ([note_w] if has_notes else [])
        pdf._table_header(all_widths, all_headers)
        pdf.set_font(_FONT_NAME, "", 7.5)

        ok = 0
        for _, d in details.iterrows():
            row_h = d_rh
            napomena = d.get("napomena", "")

            if pdf.get_y() + row_h > pdf.h - 18:
                pdf.add_page()
                pdf.set_font(_FONT_NAME, "B", 9)
                pdf.set_text_color(*_GREEN_DARK)
                pdf.cell(0, 6, f"{pj_code} — {pj_name} (nastavak)", new_x="LMARGIN", new_y="NEXT")
                pdf.ln(1)
                pdf._table_header(all_widths, all_headers)
                pdf.set_font(_FONT_NAME, "", 7.5)

            pdf._status_fill(d["status"])
            date_str = d["date"].strftime("%d.%m.%Y") if pd.notna(d["date"]) else ""
            pdf.cell(d_widths[0], row_h, date_str, border=1, fill=True)
            pdf.cell(d_widths[1], row_h, f"{d['promet']:,.2f}", border=1, fill=True, align="R")
            pdf.cell(d_widths[2], row_h, f"{d['banka']:,.2f}", border=1, fill=True, align="R")
            pdf.cell(d_widths[3], row_h, f"{d['razlika']:+,.2f}", border=1, fill=True, align="R")
            pdf.cell(d_widths[4], row_h, d["status"], border=1, fill=True, align="C")
            if has_notes:
                pdf.set_font(_FONT_NAME, "", 6)
                pdf.cell(note_w, row_h, str(napomena)[:55], border=1, fill=True)
                pdf.set_font(_FONT_NAME, "", 7.5)
            pdf.ln()

            if d["status"] in ("OK", "OK*"):
                ok += 1

        pdf.set_font(_FONT_NAME, "", 7)
        pdf.set_text_color(*_GRAY)
        pdf.cell(0, 4, f"Podudaranje: {ok}/{len(details)} dana", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(4)

    return bytes(pdf.output())
