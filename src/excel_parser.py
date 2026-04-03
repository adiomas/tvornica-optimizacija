import re
from datetime import datetime

import openpyxl
import pandas as pd


def parse_date(date_str: str) -> datetime | None:
    """Parse date string like '16.3.2026.' or '1.4.2026.' into datetime."""
    if not date_str:
        return None
    s = str(date_str).strip().rstrip(".")
    for fmt in ("%d.%m.%Y", "%d.%m.%y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def parse_excel(file) -> tuple[pd.DataFrame, dict]:
    """Parse OLAP Excel file into a DataFrame.

    Args:
        file: file path (str) or file-like object (Streamlit UploadedFile)

    Returns:
        (df, meta) where df has columns [pj_code, pj_name, date, amount]
        and meta has keys: period_start, period_end, branch_count
    """
    wb = openpyxl.load_workbook(file, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = []
    current_pj = None
    current_name = None

    for row in ws.iter_rows(min_row=8, max_row=ws.max_row, values_only=False):
        a_val = row[0].value   # Column A - PJ code
        f_val = row[5].value   # Column F - PJ name
        h_val = row[7].value   # Column H - Date
        j_val = row[9].value   # Column J - Payment type
        l_val = row[11].value  # Column L - Amount

        # New PJ section
        if a_val and str(a_val).startswith("PJ"):
            current_pj = str(a_val).strip()
            current_name = str(f_val).strip() if f_val else current_pj

        # Skip total rows and header rows
        if f_val and "Total" in str(f_val):
            continue
        if not h_val or not l_val:
            continue

        dt = parse_date(h_val) if isinstance(h_val, str) else h_val
        if dt is None:
            continue

        amount = float(l_val) if l_val else 0.0

        rows.append({
            "pj_code": current_pj,
            "pj_name": current_name,
            "date": dt,
            "amount": round(amount, 2),
        })

    wb.close()

    df = pd.DataFrame(rows)
    if df.empty:
        return df, {"period_start": None, "period_end": None, "branch_count": 0}

    df["date"] = pd.to_datetime(df["date"])

    meta = {
        "period_start": df["date"].min(),
        "period_end": df["date"].max(),
        "branch_count": df["pj_code"].nunique(),
    }
    return df, meta
